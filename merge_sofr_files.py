"""
merge_sofr_files.py
-------------------
Merges NY Fed SOFR Rate files and SOFR Index files into a single
Excel file ready for import into the SOFR Interest Calculator app.

Supported file formats (auto-detected):

  Format A — Combined file (single file with all columns):
    Effective Date | Rate Type | Rate (%) | ... | SOFR Index | ...
    Both Rate (%) and SOFR Index are extracted from this one file.
    Place in EITHER folder — the script detects it automatically.

  Format B — Separate files:
    Rate file  : DATE/Effective Date | BENCHMARK NAME/Rate Type | RATE (%)/Rate (%)
    Index file : DATE/Effective Date | BENCHMARK NAME/Rate Type | INDEX/SOFR Index

Folder structure (relative to this script):
    SOFR Rate/      ← NY Fed SOFR rate Excel files
    SOFR Index/     ← NY Fed SOFR index Excel files
    SOFR_Merged/    ← output written here (auto-created)

Usage:
    python merge_sofr_files.py

NY Fed downloads:
    https://www.newyorkfed.org/markets/reference-rates/sofr
    https://www.newyorkfed.org/markets/reference-rates/sofr-averages-and-index
"""

import sys
from pathlib import Path
from datetime import datetime

import pandas as pd


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

SCRIPT_DIR   = Path(__file__).parent
RATE_FOLDER  = SCRIPT_DIR / "SOFR Rate"
INDEX_FOLDER = SCRIPT_DIR / "SOFR Index"
OUT_FOLDER   = SCRIPT_DIR / "SOFR_Merged"
SUPPORTED    = {".xlsx", ".xls"}

# Column name variants the NY Fed has used over the years
DATE_VARIANTS  = {"date", "effective date", "effectivedate"}
RATE_VARIANTS  = {"rate (%)", "rate(%)", "sofr rate (%)", "sofrrate(%)", "rate (%)"}
INDEX_VARIANTS = {"sofr index", "sofrindex", "index"}
BM_VARIANTS    = {"rate type", "ratetype", "benchmark name", "benchmarkname",
                  "benchmark", "type"}

# Values in the benchmark/rate-type column that mean plain overnight SOFR
SOFR_BM_VALUES  = {"sofr", "overnight sofr"}
# Values that mean the SOFR Averages & Index series
SOFRAI_BM_VALUES = {"sofrai", "sofr index", "sofr compounded index",
                    "sofr averages and index"}


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _n(col: str) -> str:
    """Normalise a column name for comparison."""
    return col.strip().lower()


def find_excel_files(folder: Path) -> list[Path]:
    if not folder.exists():
        raise FileNotFoundError(
            f"Folder not found: {folder}\n"
            f"Please create the folder and place your Excel files inside."
        )
    files = [f for f in sorted(folder.iterdir()) if f.suffix.lower() in SUPPORTED]
    if not files:
        raise FileNotFoundError(
            f"No Excel files found in: {folder}"
        )
    return files


def _find_col(df: pd.DataFrame, variants: set) -> str | None:
    """Return the first column whose normalised name is in variants."""
    for col in df.columns:
        if _n(col) in variants:
            return col
    return None


def _parse_numeric(series: pd.Series) -> pd.Series:
    return (
        series.astype(str)
        .str.strip()
        .str.replace(",", ".", regex=False)
        .pipe(pd.to_numeric, errors="coerce")
    )


def _filter_sofr_rows(df: pd.DataFrame, bm_col: str,
                      keep_values: set) -> pd.DataFrame:
    """Keep only rows whose benchmark/rate-type value is in keep_values."""
    mask = df[bm_col].astype(str).str.strip().str.lower().isin(keep_values)
    filtered = df[mask].copy()
    return filtered


# ---------------------------------------------------------------------------
# File readers
# ---------------------------------------------------------------------------

def read_file(path: Path) -> dict:
    """
    Read one Excel file and return a dict with keys:
        'rates'  -> DataFrame(date, sofr_rate)   or None
        'index'  -> DataFrame(date, sofr_index)  or None
    Detects automatically whether the file contains rates, index, or both.
    """
    print(f"\n  Reading: {path.name}")

    df = pd.read_excel(path, header=0)
    df.columns = [str(c).strip() for c in df.columns]

    print(f"    Columns: {list(df.columns)}")

    date_col  = _find_col(df, DATE_VARIANTS)
    rate_col  = _find_col(df, RATE_VARIANTS)
    index_col = _find_col(df, INDEX_VARIANTS)
    bm_col    = _find_col(df, BM_VARIANTS)

    if date_col is None:
        raise ValueError(
            f"Cannot find a date column.\n"
            f"    Expected one of: {DATE_VARIANTS}\n"
            f"    Found: {list(df.columns)}"
        )

    # Parse date column
    df["_date"] = pd.to_datetime(df[date_col], errors="coerce").dt.date
    df = df.dropna(subset=["_date"])

    result = {"rates": None, "index": None}

    # ── Extract SOFR Rate (%%) ──────────────────────────────────────────────
    if rate_col:
        sub = df[["_date", rate_col, *(([bm_col]) if bm_col else [])]].copy()
        sub["_rate"] = _parse_numeric(sub[rate_col])
        sub = sub.dropna(subset=["_rate"])

        # Filter to plain SOFR rows if we have a benchmark column
        if bm_col and bm_col in sub.columns:
            filtered = _filter_sofr_rows(sub, bm_col, SOFR_BM_VALUES)
            if len(filtered) > 0:
                sub = filtered
            # else: no filter match — use all rows (single-series file)

        sub = sub[["_date", "_rate"]].rename(
            columns={"_date": "date", "_rate": "sofr_rate"}
        )
        sub = sub.drop_duplicates("date").sort_values("date")
        result["rates"] = sub
        print(f"    Rate rows  : {len(sub)}  "
              f"({sub['date'].min()} to {sub['date'].max()})")

    # ── Extract SOFR Index ─────────────────────────────────────────────────
    if index_col:
        sub = df[["_date", index_col, *(([bm_col]) if bm_col else [])]].copy()
        sub["_idx"] = _parse_numeric(sub[index_col])
        sub = sub.dropna(subset=["_idx"])

        # Filter to SOFRAI rows if benchmark column present
        if bm_col and bm_col in sub.columns:
            filtered = _filter_sofr_rows(sub, bm_col, SOFRAI_BM_VALUES)
            if len(filtered) > 0:
                sub = filtered
            # else: single-series file — use all rows

        sub = sub[["_date", "_idx"]].rename(
            columns={"_date": "date", "_idx": "sofr_index"}
        )
        sub = sub.drop_duplicates("date").sort_values("date")
        result["index"] = sub
        print(f"    Index rows : {len(sub)}  "
              f"({sub['date'].min()} to {sub['date'].max()})")

    if result["rates"] is None and result["index"] is None:
        raise ValueError(
            f"Could not extract rates or index.\n"
            f"    Rate col found   : {rate_col}\n"
            f"    Index col found  : {index_col}\n"
            f"    Columns in file  : {list(df.columns)}"
        )

    return result


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def merge_sofr_files():
    print("=" * 60)
    print("  SOFR Rate + Index File Merger")
    print("=" * 60)

    all_rates  = []
    all_index  = []

    # ── Process SOFR Rate folder ────────────────────────────────────────────
    print(f"\n[1] SOFR Rate folder: {RATE_FOLDER}")
    for path in find_excel_files(RATE_FOLDER):
        try:
            res = read_file(path)
            if res["rates"]  is not None: all_rates.append(res["rates"])
            if res["index"]  is not None: all_index.append(res["index"])
        except Exception as e:
            print(f"  WARNING: Skipping {path.name}\n    {e}")

    # ── Process SOFR Index folder ───────────────────────────────────────────
    print(f"\n[2] SOFR Index folder: {INDEX_FOLDER}")
    for path in find_excel_files(INDEX_FOLDER):
        try:
            res = read_file(path)
            if res["rates"]  is not None: all_rates.append(res["rates"])
            if res["index"]  is not None: all_index.append(res["index"])
        except Exception as e:
            print(f"  WARNING: Skipping {path.name}\n    {e}")

    # ── Validate ────────────────────────────────────────────────────────────
    if not all_rates:
        print("\nERROR: No SOFR Rate (%) data found in any file.")
        print("  Make sure your files contain a 'Rate (%)' column.")
        sys.exit(1)

    if not all_index:
        print("\nERROR: No SOFR Index data found in any file.")
        print("  Make sure your files contain a 'SOFR Index' column.")
        sys.exit(1)

    # ── Combine and deduplicate ─────────────────────────────────────────────
    rates_df = (pd.concat(all_rates, ignore_index=True)
                .drop_duplicates("date")
                .sort_values("date")
                .reset_index(drop=True))

    index_df = (pd.concat(all_index, ignore_index=True)
                .drop_duplicates("date")
                .sort_values("date")
                .reset_index(drop=True))

    print(f"\n[3] Combined totals before merge:")
    print(f"    Rate rows  : {len(rates_df)}  "
          f"({rates_df['date'].min()} to {rates_df['date'].max()})")
    print(f"    Index rows : {len(index_df)}  "
          f"({index_df['date'].min()} to {index_df['date'].max()})")

    # ── Inner join on date ──────────────────────────────────────────────────
    print("\n[4] Merging on Date (inner join)…")
    merged = pd.merge(rates_df, index_df, on="date", how="inner")
    merged = merged.sort_values("date", ascending=False).reset_index(drop=True)

    rate_only  = len(rates_df)  - len(merged)
    index_only = len(index_df) - len(merged)

    print(f"    Merged rows         : {len(merged)}")
    if rate_only:
        print(f"    Dates in rate only  : {rate_only}  (no matching index date)")
    if index_only:
        print(f"    Dates in index only : {index_only} (no matching rate date)")

    if len(merged) == 0:
        print("\nERROR: No matching dates between rate and index data.")
        print("  Ensure both files cover overlapping date ranges.")
        sys.exit(1)

    # ── Rename to match app import format ───────────────────────────────────
    merged = merged.rename(columns={
        "date":       "Date",
        "sofr_rate":  "SOFR Rate (%)",
        "sofr_index": "SOFR Index",
    })

    # Day Count Factor: Friday = 3, Mon-Thu = 1
    merged["Day Count Factor"] = merged["Date"].apply(
        lambda d: 3 if d.weekday() == 4 else 1
    )

    # ── Write output Excel ──────────────────────────────────────────────────
    OUT_FOLDER.mkdir(exist_ok=True)
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = OUT_FOLDER / f"SOFR_Rates_Merged_{ts}.xlsx"

    from openpyxl.styles import Font, PatternFill, Alignment

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        merged.to_excel(writer, index=False, sheet_name="SOFR Rates")
        ws = writer.sheets["SOFR Rates"]

        # Header styling
        hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill("solid", start_color="1B3A6B")
        hdr_align = Alignment(horizontal="center", vertical="center")
        for cell in ws[1]:
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align

        # Column widths and number formats
        col_formats = {
            "A": ("Date",             18, "MM/DD/YYYY"),
            "B": ("SOFR Rate (%)",    16, "0.0000"),
            "C": ("SOFR Index",       18, "0.00000000"),
            "D": ("Day Count Factor", 16, "0"),
        }
        for col, (_, width, fmt) in col_formats.items():
            ws.column_dimensions[col].width = width
            for row in ws.iter_rows(min_row=2,
                                    min_col=ord(col) - 64,
                                    max_col=ord(col) - 64):
                for cell in row:
                    cell.number_format = fmt

        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

    # ── Summary ─────────────────────────────────────────────────────────────
    print(f"\n{'=' * 60}")
    print(f"  Output : {out_path}")
    print(f"  Rows   : {len(merged)}")
    print(f"  Range  : {merged['Date'].min()} to {merged['Date'].max()}")
    print(f"  Columns: {list(merged.columns)}")
    print(f"{'=' * 60}")
    print("\nDone.  Import this file via:")
    print("  App → SOFR Rates (left nav) → Browse → Import Rates")

    return out_path


if __name__ == "__main__":
    try:
        merge_sofr_files()
    except FileNotFoundError as e:
        print(f"\nERROR: {e}")
        sys.exit(1)
    except Exception as e:
        import traceback
        print(f"\nUnexpected error: {e}")
        traceback.print_exc()
        sys.exit(1)
